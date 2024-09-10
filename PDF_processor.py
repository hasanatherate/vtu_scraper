import re
from pathlib import Path
import pdfplumber
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

def extract_text_from_pdf(pdf_path):
    text = ""
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text
    logging.debug(f"Extracted text from {pdf_path}:\n{text}")
    return text

def extract_info(content):
    usn_match = re.search(r'University Seat Number\s*:\s*(\w+)', content)
    name_match = re.search(r'Student Name\s*:\s*(.+)', content)
    semester_match = re.search(r'Semester\s*:\s*(\d+)', content)
    
    usn = usn_match.group(1) if usn_match else "N/A"
    name = name_match.group(1) if name_match else "N/A"
    semester = semester_match.group(1) if semester_match else "N/A"
    
    subject_pattern = r'(\w+):?\s*(.*?)\s+(\d+)\s+(\d+)\s+(\d+)\s+([PF])\s+(\d{4}-\d{2}-\d{2})'
    subjects = re.findall(subject_pattern, content)
    
    logging.debug(f"Extracted info - USN: {usn}, Name: {name}, Semester: {semester}, Subjects: {subjects}")
    
    return usn, name, semester, subjects

def process_pdf_content(content):
    usn, name, semester, subjects = extract_info(content)
    rows = []
    
    for subject in subjects:
        code, subject_name, internal, external, total, result, date = subject
        
        # Ensure the subject code follows the BVLXXX format
        if not re.match(r'BVL\d{3}', code):
            parts = subject_name.split()
            for part in parts:
                if re.match(r'BVL\d{3}', part):
                    code = part
                    subject_name = subject_name.replace(part, "").strip()
                    break
        
        # Ensure no numbers in the subject name (move numbers to subject code if necessary)
        parts = subject_name.split()
        subject_name_filtered = []
        for part in parts:
            if any(char.isdigit() for char in part):
                code = part  # Assume it's part of the subject code if it contains numbers
            else:
                subject_name_filtered.append(part)
        
        subject_name = ' '.join(subject_name_filtered).strip()
        
        rows.append([usn, name, semester, code.strip(), subject_name.strip(), internal, external, total, result, date])
    
    logging.debug(f"Processed rows: {rows}")
    return rows

def create_excel_file(data, output_file):
    wb = Workbook()
    wb.remove(wb.active)  # Remove the default sheet

    headers = ['Subject Code', 'Subject Name', 'Internal Marks', 'External Marks', 'Total Marks', 'Result', 'Date']

    for student_data in data:
        if not student_data:
            continue
        usn, name, semester = student_data[0][:3]  # Extracting USN, Name, and Semester from first row
        sheet = wb.create_sheet(title=usn)  # Use USN as sheet name

        # Styling
        header_font = Font(bold=True)
        centered_alignment = Alignment(horizontal='center', vertical='center')
        left_alignment = Alignment(horizontal='left', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Student info row
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        student_info_cell = sheet.cell(row=1, column=1, value=f"USN: {usn}, Name: {name}, Semester: {semester}")
        student_info_cell.font = Font(bold=True, size=12)
        student_info_cell.alignment = centered_alignment
        sheet.row_dimensions[1].height = 25

        # Headers
        for col_num, header in enumerate(headers, start=1):
            cell = sheet.cell(row=2, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = centered_alignment
            cell.border = border

        # Data rows
        for row_num, row in enumerate(student_data, start=3):
            for col_num, value in enumerate(row[3:], start=1):  # Start from Subject Code
                cell = sheet.cell(row=row_num, column=col_num, value=value)
                cell.alignment = centered_alignment if col_num != 2 else left_alignment  # Left-align subject names
                cell.border = border

        # Adjust column widths
        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_file)
    logging.info(f"Excel file '{output_file}' has been created with improved formatting.")

def main():
    input_folder = Path("/Users/abdul/Downloads/pdf downloads")
    output_file = input_folder / "results13.xlsx"

    if not input_folder.exists():
        logging.error(f"The directory {input_folder} does not exist.")
        return

    if not input_folder.is_dir():
        logging.error(f"{input_folder} is not a directory.")
        return
    
    all_data = {}
    
    pdf_files = list(input_folder.glob('*.pdf'))
    
    if not pdf_files:
        logging.warning(f"No PDF files found in {input_folder}")
        return
    
    for pdf_file in pdf_files:
        logging.info(f"Processing {pdf_file}")
        try:
            content = extract_text_from_pdf(pdf_file)
            rows = process_pdf_content(content)
            for row in rows:
                usn = row[0]
                if usn not in all_data:
                    all_data[usn] = []
                all_data[usn].append(row)
            logging.info(f"Successfully processed {pdf_file}")
        except Exception as e:
            logging.error(f"Error processing {pdf_file}: {str(e)}")
    
    create_excel_file(list(all_data.values()), output_file)

if __name__ == "__main__":
    main()
